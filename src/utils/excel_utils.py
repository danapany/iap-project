# utils/excel_utils.py
import pandas as pd
import streamlit as st
import re
from io import BytesIO
from datetime import datetime

class ExcelDownloadManager:
    """엑셀 다운로드 관리 클래스"""
    
    def __init__(self):
        self.debug_mode = False
    
    def extract_table_from_response(self, response_text):
        """응답 텍스트에서 표 데이터 추출 - 개선된 버전"""
        try:
            # 마크다운 표 패턴 찾기 (더 유연한 패턴)
            table_pattern = r'\|([^|]+\|)*[^|]+\|'
            
            lines = response_text.split('\n')
            table_lines = []
            in_table = False
            
            print(f"DEBUG: 전체 라인 수: {len(lines)}")  # 디버그 로그
            
            for i, line in enumerate(lines):
                line = line.strip()
                
                # 표 시작 감지 - 더 관대한 조건
                if '|' in line and len(line.split('|')) >= 3:
                    # 구분선(---|---|---) 스킵
                    if re.match(r'\|[\s\-:]+\|', line):
                        print(f"DEBUG: 구분선 스킵: {line}")
                        continue
                        
                    table_lines.append(line)
                    in_table = True
                    print(f"DEBUG: 표 라인 추가: {line}")
                    
                elif in_table and not line:
                    # 빈 줄이 나오면 표 끝
                    print(f"DEBUG: 빈 줄로 표 종료")
                    break
                elif in_table and '|' not in line:
                    # 파이프가 없는 줄이 나오면 표 끝
                    print(f"DEBUG: 파이프 없는 줄로 표 종료: {line}")
                    break
            
            print(f"DEBUG: 추출된 표 라인 수: {len(table_lines)}")
            
            if not table_lines:
                # 표를 찾지 못한 경우 더 관대한 검색
                return self._extract_table_fallback(response_text)
            
            # 표 데이터 파싱
            parsed_data = []
            headers = None
            
            for i, line in enumerate(table_lines):
                # 파이프로 분리하고 앞뒤 공백 제거
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                
                if not cells:
                    continue
                    
                if i == 0:
                    headers = cells
                    print(f"DEBUG: 헤더 추출: {headers}")
                else:
                    if len(cells) >= len(headers) - 1:  # 약간의 오차 허용
                        # 셀 수가 부족하면 빈 값으로 채움
                        while len(cells) < len(headers):
                            cells.append("")
                        parsed_data.append(cells[:len(headers)])
                        print(f"DEBUG: 데이터 행 추가: {cells[:len(headers)]}")
            
            if headers and parsed_data:
                result = {'headers': headers, 'data': parsed_data}
                print(f"DEBUG: 파싱 성공 - 헤더: {len(headers)}, 데이터: {len(parsed_data)}")
                return result
            
            return None
            
        except Exception as e:
            print(f"DEBUG: Table extraction error: {e}")
            return None

    def _extract_table_fallback(self, response_text):
        """표 추출 실패 시 폴백 방법"""
        try:
            # 리스트 형태에서 표 데이터 추출 시도
            lines = response_text.split('\n')
            
            # "장애 ID", "서비스명" 등의 키워드가 포함된 라인 찾기
            table_keywords = ['장애 ID', '서비스명', '장애등급', '발생일자']
            
            potential_headers = []
            potential_data = []
            
            for line in lines:
                line = line.strip()
                
                # 표 헤더 후보 찾기
                if any(keyword in line for keyword in table_keywords) and '|' in line:
                    potential_headers = [cell.strip() for cell in line.split('|') if cell.strip()]
                    print(f"DEBUG: Fallback 헤더 발견: {potential_headers}")
                    continue
                    
                # 데이터 행 후보 찾기 (INM으로 시작하는 장애 ID 포함)
                if 'INM' in line and '|' in line:
                    cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                    if len(cells) >= 3:  # 최소 3개 이상의 셀
                        potential_data.append(cells)
                        print(f"DEBUG: Fallback 데이터 발견: {cells}")
            
            if potential_headers and potential_data:
                # 헤더와 데이터 길이 맞추기
                max_cols = max(len(potential_headers), max(len(row) for row in potential_data))
                
                while len(potential_headers) < max_cols:
                    potential_headers.append(f"컬럼{len(potential_headers)+1}")
                    
                for row in potential_data:
                    while len(row) < max_cols:
                        row.append("")
                
                return {'headers': potential_headers, 'data': potential_data}
            
            return None
            
        except Exception as e:
            print(f"DEBUG: Fallback extraction error: {e}")
            return None

    def create_excel_dataframe(self, table_data):
        """표 데이터에서 pandas DataFrame 생성"""
        try:
            if not table_data or not table_data.get('headers') or not table_data.get('data'):
                return None
            
            df = pd.DataFrame(table_data['data'], columns=table_data['headers'])
            
            # 데이터 타입 정리
            for col in df.columns:
                if '시간' in col and '분' in col:
                    # 장애시간(분) 컬럼을 숫자로 변환
                    df[col] = pd.to_numeric(df[col].str.replace('분', ''), errors='coerce')
                elif '일자' in col or '날짜' in col:
                    # 날짜 컬럼 처리
                    df[col] = pd.to_datetime(df[col], errors='coerce')
            
            return df
            
        except Exception as e:
            if self.debug_mode:
                print(f"DEBUG: DataFrame creation error: {e}")
            return None
    
    def generate_excel_file(self, df, filename_prefix="장애내역"):
        """DataFrame을 엑셀 파일로 변환"""
        try:
            if df is None or df.empty:
                return None
            
            # 현재 시간을 파일명에 포함
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            # BytesIO 객체에 엑셀 파일 생성
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='장애내역', index=False)
                
                # 워크시트 스타일 적용
                worksheet = writer.sheets['장애내역']
                
                # 헤더 스타일 적용
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_alignment = Alignment(horizontal='center', vertical='center')
                
                # 헤더 행에 스타일 적용
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = center_alignment
                
                # 데이터 행에 테두리 적용
                for row_num in range(2, len(df) + 2):
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # 컬럼 너비 자동 조정
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            excel_data = output.getvalue()
            return excel_data, filename
            
        except Exception as e:
            if self.debug_mode:
                print(f"DEBUG: Excel generation error: {e}")
            return None, None
    
    def display_download_button(self, response_text, query_type="inquiry"):
        """엑셀 다운로드 버튼 표시 - 개선된 버전"""
        try:
            print(f"DEBUG: 다운로드 버튼 표시 시작 - query_type: {query_type}")
            
            # INQUIRY 타입이 아니면 다운로드 버튼 표시하지 않음
            if query_type.lower() != 'inquiry':
                print(f"DEBUG: INQUIRY 타입이 아니므로 다운로드 버튼 표시하지 않음")
                return False
            
            # 응답에서 표 데이터 추출
            table_data = self.extract_table_from_response(response_text)
            
            if not table_data:
                print(f"DEBUG: 표 데이터를 찾을 수 없음")
                # 강제로 간단한 표 생성 시도
                if "장애 ID" in response_text and "서비스명" in response_text:
                    table_data = self._create_simple_table_from_response(response_text)
            
            if not table_data:
                print(f"DEBUG: 표 데이터 생성 실패")
                return False
            
            print(f"DEBUG: 표 데이터 추출 성공: {len(table_data.get('data', []))} 행")
            
            # DataFrame 생성
            df = self.create_excel_dataframe(table_data)
            
            if df is None or df.empty:
                print(f"DEBUG: DataFrame 생성 실패")
                return False
            
            print(f"DEBUG: DataFrame 생성 성공: {len(df)} 행")
            
            # 엑셀 파일 생성
            excel_data, filename = self.generate_excel_file(df)
            
            if excel_data is None:
                print(f"DEBUG: 엑셀 파일 생성 실패")
                return False
            
            print(f"DEBUG: 엑셀 파일 생성 성공: {filename}")
            
            # 다운로드 버튼 표시
            st.markdown("---")
            st.markdown("### 📊 엑셀 다운로드")
            
            col1, col2 = st.columns([1, 3])
            
            with col1:
                st.download_button(
                    label="📥 엑셀 파일 다운로드",
                    data=excel_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="표 형태의 장애 내역을 엑셀 파일로 다운로드합니다."
                )
            
            with col2:
                st.info(f"총 {len(df)}건의 장애 내역이 포함된 엑셀 파일을 다운로드할 수 있습니다.")
            
            # 미리보기 표시
            with st.expander("📋 다운로드될 데이터 미리보기"):
                st.dataframe(df, use_container_width=True)
            
            print(f"DEBUG: 다운로드 버튼 표시 완료")
            return True
            
        except Exception as e:
        print(f"DEBUG: Download button display error: {e}")
        import traceback
        traceback.print_exc()
        
        # 에러 발생 시에도 기본 다운로드 옵션 제공
        st.markdown("---")
        st.markdown("### 📊 엑셀 다운로드")
        st.error("엑셀 다운로드 기능에 오류가 발생했습니다. 데이터를 복사하여 엑셀에 붙여넣기 해주세요.")
        return False
    
    def _create_simple_table_from_response(self, response_text):
        """응답에서 간단한 표 데이터 생성"""
        try:
            # 기본 헤더 설정
            headers = ["장애 ID", "서비스명", "장애등급", "발생일자", "시간대", "요일", "장애시간(분)", "장애현상", "장애원인", "담당부서"]
            data = []
            
            lines = response_text.split('\n')
            current_record = {}
            
            for line in lines:
                line = line.strip()
                
                # 장애 ID로 시작하는 새로운 레코드
                if line.startswith('장애 ID:') or 'INM' in line:
                    if current_record:
                        # 이전 레코드를 데이터에 추가
                        row = [current_record.get(header, "") for header in headers]
                        data.append(row)
                    current_record = {}
                    
                    # 장애 ID 추출
                    if 'INM' in line:
                        import re
                        match = re.search(r'INM\d+', line)
                        if match:
                            current_record["장애 ID"] = match.group()
                
                # 각 필드 추출
                elif '서비스명:' in line:
                    current_record["서비스명"] = line.split('서비스명:')[1].strip()
                elif '장애등급:' in line:
                    current_record["장애등급"] = line.split('장애등급:')[1].strip()
                elif '발생일자:' in line:
                    current_record["발생일자"] = line.split('발생일자:')[1].strip()
                elif '발생시간대:' in line:
                    current_record["시간대"] = line.split('발생시간대:')[1].strip()
                elif '발생요일:' in line:
                    current_record["요일"] = line.split('발생요일:')[1].strip()
                elif '장애시간:' in line:
                    current_record["장애시간(분)"] = line.split('장애시간:')[1].replace('분', '').strip()
                elif '장애현상:' in line:
                    current_record["장애현상"] = line.split('장애현상:')[1].strip()
                elif '장애원인:' in line:
                    current_record["장애원인"] = line.split('장애원인:')[1].strip()
                elif '담당부서:' in line:
                    current_record["담당부서"] = line.split('담당부서:')[1].strip()
            
            # 마지막 레코드 추가
            if current_record:
                row = [current_record.get(header, "") for header in headers]
                data.append(row)
            
            if data:
                print(f"DEBUG: 간단한 표 생성 성공: {len(data)} 행")
                return {'headers': headers, 'data': data}
            
            return None
            
        except Exception as e:
            print(f"DEBUG: Simple table creation error: {e}")
            return None
            print(f"DEBUG: Simple table creation error: {e}")
            return None    

    def extract_download_info_from_query(self, query):
        """쿼리에서 다운로드 관련 정보 추출"""
        try:
            service_patterns = [
                r'([A-Za-z가-힣][A-Za-z0-9가-힣_\-/\+\(\)\s]*[A-Za-z0-9가-힣_\-/\+\)])\s+(?:내역|목록|리스트)',
                r'([A-Za-z가-힣][A-Za-z0-9가-힣_\-/\+\(\)\s]*[A-Za-z0-9가-힣_\-/\+\)])\s+(?:장애|문제)',
            ]
            
            service_name = None
            for pattern in service_patterns:
                if matches := re.findall(pattern, query, re.IGNORECASE):
                    service_name = matches[0].strip()
                    break
            
            # 시간 정보 추출
            time_info = []
            if '야간' in query:
                time_info.append('야간')
            elif '주간' in query:
                time_info.append('주간')
            
            # 연도 추출
            year_match = re.search(r'\b(202[0-9]|201[0-9])\b', query)
            year = year_match.group(1) if year_match else None
            
            filename_parts = []
            if service_name:
                filename_parts.append(service_name)
            if year:
                filename_parts.append(f"{year}년")
            if time_info:
                filename_parts.extend(time_info)
            filename_parts.append("장애내역")
            
            return "_".join(filename_parts) if filename_parts else "장애내역"
            
        except Exception as e:
            if self.debug_mode:
                print(f"DEBUG: Download info extraction error: {e}")
            return "장애내역"